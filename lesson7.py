"""
接口自动化的步骤：
1.excel准备测试用例，代码自动读取测试数据
2.发送接口请求，得到响应信息
3.断言：实际结果 与 预期结果比较 —— 通过/不通过
4.写入通过不通过——excel
"""
import openpyxl
import requests


# 读取excel
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 加载工作蒲 —— 文档名字
    sheet = wb[sheetname]              # 获取表单
    max_row = sheet.max_row             # 获取最大行数
    max_column = sheet.max_column       # 获取最大列数
    case_list = []  # 保存读取的数据
    for row in range(2, max_row+1, 1):
        dict1 = dict(                   # 用字典类型保存起来
        cate_id = sheet.cell(row, column=1).value,
        url = sheet.cell(row, column=5).value,    # 获取单元格内容,sheet.cell(row, column=1)得到的是单元格位置标识如A2
        data = sheet.cell(row, column=6).value,
        expected = sheet.cell(row, column=7).value
        )
        #print(dict1)
        case_list.append(dict1)             # 再次保存在列表里面，没次循环都添加在数组里面
    #print(case_list)
    return case_list

# 写入结果
def write_data(filename, sheetname, row, column, final_relst):
    wb = openpyxl.load_workbook(filename)  # 加载工作蒲 —— 文档名字
    sheet = wb[sheetname]  # 获取表单
    sheet.cell(row, column).value = final_relst
    wb.save(filename)

# 请求函数
def api_fun(url,data):
    reg = requests.post(url=url, json=data, headers={'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'})
    response = reg.json()  # 响应正文
    return response


cases = read_data('text_api_register.xlsx', 'register')  # 读取出excel的数据
for case in cases:
    cate_id = case.get('cate_id')
    url = case.get('url')
    data = eval(case.get('data'))   # case.get('data')得到的是一个字符串类型的数据，eval(）可以去掉字符串的引号，恢复原有数据类型
    expected = eval(case.get('expected'))
    #print(cate_id, url, data, expected)
    response = api_fun(url, data)   # 发送请求
    if ((response.get('code')==expected.get('code')) and (response.get('msg')==expected.get('msg'))):  # 与预期结果判断，断言
        print('ture')
        write_data('text_api_register.xlsx', 'register', cate_id+1, 8, 'ture')  # 写入数据
        write_data('text_api_register.xlsx', 'register', cate_id+1, 9, response.get('msg'))
    else:
        print('false')
        write_data('text_api_register.xlsx', 'register', cate_id + 1, 8, 'false')
        write_data('text_api_register.xlsx', 'register', cate_id+1, 9, response.get('msg'))
    # print(expected.get("code"))
